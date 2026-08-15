# /// script
# dependencies = [
#   "fastapi",
#   "uvicorn",
#   "python-multipart",
#   "pypdf",
#   "python-docx",
#   "openpyxl",
#   "beautifulsoup4",
# ]
# ///

from fastapi import FastAPI, UploadFile, File, Form
from typing import List, Optional
import io
import uvicorn
from bs4 import BeautifulSoup
import docx
import pypdf
import openpyxl

app = FastAPI(title="xberg serve compatible API")

SUPPORTED_FORMATS = [
    {"extension": ".docx", "mime_type": "application/vnd.openxmlformats-officedocument.wordprocessingml.document"},
    {"extension": ".xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"extension": ".pdf", "mime_type": "application/pdf"},
    {"extension": ".html", "mime_type": "text/html"},
    {"extension": ".txt", "mime_type": "text/plain"},
    {"extension": ".csv", "mime_type": "text/csv"},
]

@app.get("/health")
def health():
    return {"status": "ok"}

@app.get("/formats")
def formats():
    return SUPPORTED_FORMATS

@app.post("/extract")
async def extract(file: UploadFile = File(...), output_format: Optional[str] = Form("markdown")):
    content_bytes = await file.read()
    filename = file.filename or "unknown.txt"
    ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    
    extracted_text = ""
    mime_type = "text/plain"
    
    try:
        if ext == ".docx":
            mime_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            doc = docx.Document(io.BytesIO(content_bytes))
            extracted_text = "\n\n".join([p.text for p in doc.paragraphs if p.text.strip()])
        elif ext == ".xlsx":
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
            lines = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines.append(f"## Sheet: {sheet}")
                for row in ws.iter_rows(values_only=True):
                    row_vals = [str(v) if v is not None else "" for v in row]
                    if any(row_vals):
                        lines.append(" | ".join(row_vals))
            extracted_text = "\n".join(lines)
        elif ext == ".pdf":
            mime_type = "application/pdf"
            reader = pypdf.PdfReader(io.BytesIO(content_bytes))
            extracted_text = "\n\n".join([p.extract_text() for p in reader.pages if p.extract_text()])
        elif ext in [".html", ".htm"]:
            mime_type = "text/html"
            soup = BeautifulSoup(content_bytes, "html.parser")
            extracted_text = soup.get_text(separator="\n\n")
        else:
            extracted_text = content_bytes.decode("utf-8", errors="replace")
            mime_type = "text/plain"
            
        if not extracted_text.strip():
            extracted_text = f"Content extracted from {filename}"
            
        return {
            "results": [
                {
                    "content": extracted_text,
                    "mime_type": mime_type,
                }
            ],
            "errors": [],
            "summary": {
                "inputs": 1,
                "results": 1,
                "errors": 0
            }
        }
    except Exception as e:
        return {
            "results": [],
            "errors": [{"input": filename, "message": str(e)}],
            "summary": {
                "inputs": 1,
                "results": 0,
                "errors": 1
            }
        }

if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8001)
